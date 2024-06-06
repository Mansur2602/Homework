import openpyxl



def read_numbers_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    numbers = []
    for row in sheet.iter_rows(values_only=True):
        numbers.append(row)
    return numbers


numbers_file1 = read_numbers_from_excel('file1.xlsx')
numbers_file2 = read_numbers_from_excel('file2.xlsx')
numbers_file3 = read_numbers_from_excel('file3.xlsx')


matrix = []
for row in numbers_file1:
    matrix.append(row)
for row in numbers_file2:
    matrix.append(row)
for row in numbers_file3:
    matrix.append(row)
print(matrix)

flipped_matrix = [tuple(sorted(row, reverse=True)) for row in reversed(matrix)]
for row in flipped_matrix:
    print(row)



from openpyxl.styles import Font, Border, Side


wb = openpyxl.Workbook()
sheet = wb.active


for row_index, row_value in enumerate(matrix, start=1):
    for col_index, cell_value in enumerate(row_value, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_value)


font = Font(name='Arial', size=12, bold=True)
border = Border(left=Side(border_style='thin'), 
                right=Side(border_style='thin'), 
                top=Side(border_style='thin'), 
                bottom=Side(border_style='thin'))

for row in sheet.iter_rows():
    for cell in row:
        cell.font = font
        cell.border = border

# Сохраняем файл
wb.save('Matrix.xlsx')
